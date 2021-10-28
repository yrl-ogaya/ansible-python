import openpyxl
import yaml
import logging
import sys
from rich.logging import RichHandler

# ストリームハンドラの設定とファイルハンドラの設定
rich_handler: RichHandler = RichHandler(rich_tracebacks=True)
rich_handler.setLevel(logging.INFO)
rich_handler.setFormatter(logging.Formatter("%(message)s"))
logging.basicConfig(level=logging.INFO, handlers=[rich_handler])

# Logger名の指定。
logger = logging.getLogger(__name__)

# Excel ファイルを指定する
wbook = openpyxl.load_workbook("AnsiblePythonLab.xlsx")

# Excel ファイルのシートリストを取得する
sheets = wbook.sheetnames

# sheet の枚数分繰り返す
for sheet in sheets:
  # 出力ファイルの指定
  yaml_output = open("host_vars/"+ sheet + ".yaml", 'w')

  logger.info("-- Processing : " + sheet + " sheet --")
  sheet = wbook[sheet]
  hostname = sheet["B2"]
  ip = sheet["B3"]
  cider = sheet["B4"]
  gateway = sheet["B5"]
  dns = sheet["B6"]
  packages = list()

  # インストールパッケージリストの取得
  row_no = 11
  col_no = 1
  max_row_no = 23
  while row_no < max_row_no:
      cell_value = sheet.cell(row=row_no,column=col_no).value
      if cell_value == None:
          break
      packages.append(cell_value)
      row_no = row_no + 1

  # yaml 出力データの整形
  out_data = {'hostanme': hostname.value,
              'ip': ip.value,
              'cider': cider.value,
              'gateway': gateway.value,
              'dns': dns.value,
              'packa': packages
             }
  
  # 各シートの情報を yaml に変換して出力
  yaml_output.write(yaml.dump(out_data))
  yaml_output.close()
  
  # yaml出力内容の出力
  logger.info("hostname: " + str(hostname.value))
  logger.info("ip : " + str(ip.value))
  logger.info("cider: " + str(cider.value))
  logger.info("gateway: " + str(gateway.value))
  logger.info("dns: " + str(dns.value))
  logger.info("packages: " + str(packages))
 