# import 宣言
import openpyxl
import yaml
import sys

# main 
def main():
  # Excel ファイルを指定する
  wbook = openpyxl.load_workbook("AnsibleSample.xlsx")
  
  # Excel ファイルのシートリストを取得する
  sheets = wbook.sheetnames
  
  # sheet の枚数分繰り返す
  for sheetname in sheets:

    sheet = wbook[sheetname] # 処理するシートの取得
    hostname = sheet["B3"]
    ip = sheet["B4"]
    cider = sheet["B5"]
    gateway = sheet["B6"]
    dns = sheet["B7"]
    packages = list()
  
    # インストールパッケージリストの取得
    packages = getList(sheet,row_no=12,col_no=1,max_row_no=17) 
    # 起動/有効化サービスリストの取得
    services = getList(sheet,row_no=12,col_no=4,max_row_no=17) 
    # セキュリティファイアウォール(インバウンド許可リスト)の取得
    firewalls = getList(sheet,row_no=20,col_no=2,max_row_no=25)
    # 追加ユーザリストの取得
    users = getList(sheet,row_no=29,col_no=1,max_row_no=36)
     # 追加ユーザリストのコメント取得
    users_comment = getList(sheet,row_no=29,col_no=2,max_row_no=36)
    # yaml 変換用複合ユーザリストの作成
    users_list = list()
    for (user,comment) in zip(users,users_comment):
      user_tmp = {'name': user,'comment': comment}
      users_list.append(user_tmp)

    # yaml 出力データの整形
    out_data = {'hostanme': hostname.value,
                'ip': ip.value,
                'cider': cider.value,
                'gateway': gateway.value,
                'dns': dns.value,
                'packages': packages,
                'services': services,
                'firewalls': firewalls,
                'users': users_list
               }

    # 出力ファイルの指定
    yaml_output = open("host_vars/"+ sheetname + ".yaml", 'w') 

      # YAML へ変換
    yout = yaml.dump(out_data)
    
    # 各シートの情報を yaml に変換して出力
    yaml_output.write(yout)
    yaml_output.close()

  # man() 正常終了
  return 0

# シートリストの取得関数
def getList(sheet,row_no,col_no,max_row_no):
  convlist = list()

  # シートリストの処理
  while row_no < max_row_no:
      cell_value = sheet.cell(row=row_no,column=col_no).value
      if cell_value == None:
          row_no = row_no + 1
          continue
      convlist.append(cell_value)
      row_no = row_no + 1

  return convlist

if __name__ == '__main__':
    main() 
